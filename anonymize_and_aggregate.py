#!/bin/python3.9
"""

*******************
*** PLEASE READ ***
*******************

This module anonymizes and aggregates the peer critiques in the given directory.

Written by Akash Jain for CHE-402-SPRING-2020,
modified by Sam Hoover for CHE-401-FALL-2022.

Run code with `python3 anonymize_and_aggregate.py --path <path to folder with submissions>`.

************
*** DEMO ***
************

Using the below file structure as an example,

=======================================================================
+---CHEM-ENG 401/402
|   +---anonymize_and_aggregate (this directory)
|   |       anonymize_and_aggregate.py
|   |       requirements.txt
|   |
|   +---peer-evaluations (where to store all the peer critiques)
|   |   +---section-01
|   |   |   +---round-01
|   |   |   |   +---progress-reports
|   |   |   |   |   +---reviewer_01
|   |   |   |   |   |       review_01.xlsx
|   |   |   |   |   |       review_02.xlsx
|   |   |   |   |   |       review_03.xlsx
|   |   |   |   |   +---reviewer_02
|   |   |   |   |   |       review_01.xlsx
|   |   |   |   |   |       review_02.xlsx
|   |   |   |   |   |       review_03.xlsx
|   :   :   :   :   :
|   :   :   :   :   :
|   |   |   |   +---proposals
|   |   |   |   |   +---reviewer_01
|   |   |   |   |   |       review_01.xlsx
|   |   |   |   |   |       review_02.xlsx
|   |   |   |   |   |       review_03.xlsx
|   |   |   |   |   +---reviewer_02
|   |   |   |   |   |       review_01.xlsx
|   |   |   |   |   |       review_02.xlsx
|   |   |   |   |   |       review_03.xlsx
:   :   :   :   :   :
:   :   :   :   :   :
=======================================================================

run `python3 anonymize_and_aggregate.py --path ../peer-evaluations/section-01/round-01/proposals`
from the `anonymize-and-aggregate` directory to anonymize and aggregate the Section 1
proposal presentation peer critiques. All of the anonymized copies will be stored in the
../peer-evaluations/section-01/round-01/proposals/graded-copies` directory.

Make sure to install the required packages. Run `pip install -r requirements.txt` if unsure.
"""
import os
import shutil
import argparse
import openpyxl


class Anonymizer():
    def __init__(self, path=''):
        self.path = path

    def get_path(self):
        """
        parse the argument from the command line.

        returns:
        - path ([string]): path to directory with submissions
        """
        parser = argparse.ArgumentParser()
        parser.add_argument(
            '-p', '--path', action='store',
            help='path to directory with peer reviews'
        )

        return parser.parse_args().path


    def anonymize_reviews(self, path):
        """
        Remove the name of reviewer and save graded files in .xlxs format.

        Arguments:
        - path ([string]): path to directory with submissions
        """
        save_path = os.path.join(path, 'graded_copies')
        if os.path.exists(save_path):  # check if graded_copies already exists
            print('\ngraded-copies directory already exists, removing and starting fresh.')
            shutil.rmtree(save_path)  # if exists, delete

        os.mkdir(save_path)  # create directory for graded copies

        count = 1
        reviewers = [x for x in os.listdir(path) if x not in ['graded_copies', '.DS_Store']]
        for rev in reviewers:
            for xls in os.listdir(os.path.join(path, rev)):  # reviews

                xls = os.path.join(path, rev, xls)  # create path to spreadsheet
                wb_obj = openpyxl.load_workbook(xls, data_only=True)  # load spreadsheet
                wb_obj.worksheets[0]['D6'] = ''  # remove reviewer name

                sheet_obj = wb_obj.active
                exp = sheet_obj.cell(row=14, column=4).value
                speaker = sheet_obj.cell(row=10, column=4).value
                if speaker is None:  # if empty sheet, skip
                    print(f'An exception occured in review #{count:02}')
                    break

                nem = f'{count:02}_{speaker}_{exp}_graded.xlsx'.lower().replace(' ', '-')
                wb_obj.save(filename=os.path.join(save_path, nem))
                
                count += 1

    def aggregate(self, path):
        print(set([x.lower().split('_')[1] for x in os.listdir(os.path.join(path, 'graded_copies'))]))


def main():
    """
    Get path and then anonymize and aggregate peer critiques.
    """
    anonymizer = Anonymizer()
    path = anonymizer.get_path()
    anonymizer.anonymize_reviews(path)
    anonymizer.aggregate(path)


if __name__ == '__main__':
    main()
    print('\nDone!')
