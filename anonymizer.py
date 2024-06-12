#!/bin/python3.9
"""
*******************
*** PLEASE READ ***
*******************

This module anonymizes, aggregates, and distributes the peer critiques 
in the given directory.

Written by Akash Jain for ChE 402 in Spring 2020,
modified by Sam Hoover for ChE 401 in Fall 2022 & Fall 2023.

Run code with 
`python anonymizer.py --path <path to folder with submissions>`.

************
*** DEMO ***
************

Using the below file structure as an example,

=======================================================================
+---CHEM-ENG 401/402
|   +---anonymize_and_aggregate (this directory)
|   |       anonymize_and_aggregate.py
|   |       requirements.txt
|   |
|   +---peer-evaluations (where to store all the peer critiques)
|   |   +---section-01
|   |   |   +---round-01
|   |   |   |   +---progress-reports
|   |   |   |   |   +---reviewers
|   |   |   |   |   |   +---reviewer_01
|   |   |   |   |   |   |       review_01.xlsx
|   |   |   |   |   |   |       review_02.xlsx
|   |   |   |   |   |   |       review_03.xlsx
|   |   |   |   |   +---speakers
|   |   |   |   |   |   +---speaker_01
|   |   |   |   |   |   |       anonymized_review_01.xlsx
|   |   |   |   |   |   |       anonymized_review_02.xlsx
|   |   |   |   |   |   |       anonymized_review_03.xlsx
|   :   :   :   :   :
|   :   :   :   :   :
|   |   |   |   +---proposals
|   |   |   |   |   +---reviewers
|   |   |   |   |   |   +---reviewer_01
|   |   |   |   |   |   |       review_01.xlsx
|   |   |   |   |   |   |       review_02.xlsx
|   |   |   |   |   |   |       review_03.xlsx
|   |   |   |   |   +---speakers
|   |   |   |   |   |   +---speaker_01
|   |   |   |   |   |   |       anonymized_review_01.xlsx
|   |   |   |   |   |   |       anonymized_review_02.xlsx
|   |   |   |   |   |   |       anonymized_review_03.xlsx
:   :   :   :   :   :
:   :   :   :   :   :
=======================================================================

run `python anonymizer.py --path 
../peer-evaluations/section-01/round-01/proposals`
from the `anonymize-and-aggregate` directory to anonymize and aggregate 
the Section 1 proposal presentation peer critiques. All of the 
anonymized copies will be stored in the
../peer-evaluations/section-01/round-01/proposals/speakers` directory.

Make sure to install the required packages. Run 
`pip install -r requirements.txt` if unsure.
"""
# pyright: basic

import argparse
import os
import smtplib
from email.message import EmailMessage

import openpyxl
import pandas as pd

# TODO:
# match names based on similarity rather than manually edit names
# remove openpyxl dependencies and just use pandas


class Anonymizer:
    def __init__(self, src: str = ".", dst: str = ".", pwd: str = "") -> None:
        self.src = src
        self.dst = dst
        self.pwd = pwd

    def get_args(self) -> argparse.Namespace:
        """
        parse the argument from the command line.

        returns:
        - path ([string]): path to directory with submissions
        """
        parser: argparse.ArgumentParser = argparse.ArgumentParser()
        parser.add_argument(
            "-s", "--source", action="store", help="path to directory with peer reviews"
        )

        return parser.parse_args()

    def anonymize_reviews(self, src: str) -> None:
        """
        Remove the name of reviewer and save graded files
        in .xlxs format.

        Arguments:
        - src ([string]): path to directory with submissions
        """
        print("Begin anonymizing peer critiques...")

        dst: str = os.path.join(src, "speakers")

        # check if graded_copies already exists
        if not os.path.exists(path=dst):

            # create directory for anonymized copies
            os.mkdir(path=dst)

            count = 1
            reviewers: list[str] = [x for x in os.listdir(src) if x.endswith("_file")]

            # cycle through reviewers
            for rev in reviewers:
                # cycle through reviews
                for xls in os.listdir(os.path.join(src, rev)):

                    # create path to spreadsheet
                    xls: str = os.path.join(src, rev, xls)

                    # load spreadsheet
                    wb_obj: openpyxl.Workbook = openpyxl.load_workbook(
                        filename=xls, data_only=True
                    )

                    # remove reviewer name
                    wb_obj.worksheets[0]["D6"] = ""

                    sheet_obj = wb_obj.active

                    # extract, transform to lower case, and replace
                    # space with hyphen if cells not empty, else skip
                    try:
                        sp: str = (
                            sheet_obj.cell(row=10, column=4)
                            .value.lower()
                            .strip()
                            .replace(" ", "-")
                        )
                        exp: str = (
                            sheet_obj.cell(row=14, column=4)
                            .value.lower()
                            .strip()
                            .replace(" ", "-")
                        )
                    except ValueError as e:  # else, skip
                        print(
                            f"{e}: An exception occured in review "
                            f"#{count:02} -> {rev}."
                        )
                        break

                    # anonymized copy file name
                    nem: str = f"{count:02}_{sp}_{exp}_anonymized_copy.xlsx"

                    # check if folder exists for speaker, if not create
                    if not os.path.exists(os.path.join(dst, sp)):
                        os.mkdir(os.path.join(dst, sp))

                    # save anonymized copy to folder for speaker
                    wb_obj.save(filename=os.path.join(dst, sp, nem))

                    count += 1

            print("Done!\n")

        else:
            print("Anonymized peer critiques already exists, " "stopping...\n")

    def send_emails(self, sp_dir: str, pwd: str) -> None:
        print("Begin peer critique distribution...")

        # get peer critique type
        rev_type: str = sp_dir.split("/")[2].split("_")[0]

        # get round number
        round_num: str = sp_dir.split("/")[2].split("_")[1]

        df: pd.DataFrame = pd.read_csv("student_email_list.csv")
        for sp in os.listdir(sp_dir):
            # get speaker's email
            try:
                sp_email = df.loc[sp == df["name"], "Email address"].values[0]
            except ValueError as e:  # ask if speaker's email not found
                sp_email = input(
                    f"{e}: Could not pull the email address for {sp}, "
                    f"enter their email address: "
                )

            msg: EmailMessage = EmailMessage()
            msg["Subject"] = (
                f"Peer critiques for round {round_num} " f"{rev_type} presentation"
            )
            msg["From"] = "samuelhoover@umass.edu"
            msg["To"] = sp_email

            for f in os.listdir(os.path.join(sp_dir, sp)):
                with open(os.path.join(sp_dir, sp, f), "rb") as atchmt:
                    msg.add_attachment(
                        atchmt.read(),
                        maintype="application",
                        subtype="xlsx",
                        filename=f,
                    )

            with smtplib.SMTP("smtp.gmail.com", 587) as smtp_server:
                smtp_server.starttls()
                smtp_server.login(msg["From"], pwd)
                smtp_server.sendmail(msg["From"], msg["To"], msg.as_string())

            print(f"Email to {sp_email} sent!")

        print("Done!\n")


def create_email_list(src_csv: str) -> None:
    """
    Create readable CSV file from downloaded student list.

    To download student list, go to the "Participants" tab on the
    Moonami course page. Click on the box to select all particpants
    then scroll to the bottom of the list and download table as as .csv
    file from the "With selected users..." dropdown menu.

    Args:
        src_csv ([string]): path to .csv file of student list
    """
    # read in CSV, delete non-students, and "Groups" column
    df: pd.DataFrame = (
        pd.read_csv(src_csv).dropna().reset_index().drop(columns=["index", "Groups"])
    )

    # create full name in style "forename-surname"
    df["name"] = df[["First name", "Last name"]].apply("-".join, axis=1).str.lower()

    # save new table as CSV
    df.to_csv("student_email_list.csv")

    # delete original CSV file
    os.remove(src_csv)

    print("Created student email list\n")


def main() -> None:
    """
    Get source and destination paths and then anonymize and aggregate
    peer critiques.
    """
    if not os.path.isfile("student_email_list.csv"):
        src: str = input("Please enter the full file name for the student list: ")
        create_email_list(src_csv=src)

    anonymizer: Anonymizer = Anonymizer()
    args: argparse.Namespace = anonymizer.get_args()
    anonymizer.anonymize_reviews(src=args.source)
    input(
        "Verify spelling of speakers' names of folders are correct, "
        "ensure reviewers names are removed, and make sure to "
        "delete any empty folders, press ENTER to continue: \n"
    )
    anonymizer.send_emails(
        sp_dir=os.path.join(args.source, "speakers"),
        pwd=open("app.pwd").read().rsplit()[0],
    )


if __name__ == "__main__":
    main()
